import openpyxl as xl
from datetime import datetime, timedelta, date
from prettytable import PrettyTable
import calendar
import os

path = os.getcwd()+'/'
#path = 'C:/Users/aizat/Documents/project/'

def log_in(user):
    wb = xl.load_workbook(path + user)
    sh = wb['Logins']
    login = input('Enter your login: >>> ')
    password = input('Enter your password: >>> ')
    for lg, pw in sh['B2':'C'+str(sh.max_row)]:
        if lg.value == login and pw.value == password:
            print('You are successfully logged in!'.center(104))
            break
    else:
        while lg.value != login or pw.value != password:
            print('Wrong login or password, please try again'.center(104))
            login = input('Enter your login: >>> ')
            password = input('Enter your password: >>> ')
            for lg, pw in sh['B2':'C'+str(sh.max_row)]:
                if lg.value == login and pw.value == password:
                    print('You are successfully logged in!'.center(104))
                    break
    return login

def reg_in(user):
    wb = xl.load_workbook(path + user)
    sh = wb['Logins']
    name = input('Enter your name and surname: >>> ')
    for i in sh['A']:
        if name == i.value:
            print('You already have an account, please log in'.center(104))
            login = log_in(user)
            break
    else:
        login = input('Create your login: >>> ')
        password = input('Create your password: >>> ')
        c = 1
        while sh.cell(row=c, column=1).value != None:
            c += 1
        for nm, lg, pw in sh['A'+str(c):'C'+str(c)]:
            nm.value, lg.value, pw.value = name, login, password
        print('You have successfully registered'.center(104))
        wb.save(user)
    return login

def define_user(login, user):
    wb = xl.load_workbook(path + user)
    sh = wb['Logins']
    for nm, lg in zip(sh['A'], sh['B']):
        if login == lg.value:
            return lg.row, nm.value

def start(account):
    print('Already have an account?'.center(104))
    while True:
        is_acc = input('(y/n) >>> ').strip().lower()
        if is_acc == 'y':
            login = log_in(account)
            break
        elif is_acc == 'n':
            login = reg_in(account)
            break
        else:
            print('Wrong command'.center(104))
    row, name = define_user(login, account)
    print(f'Greetings {name}!'.center(104))
    print('Dial the menu number to work with the program'.center(104))
    return row, name

def med_hist_header():
    print('-'*104)
    print('| {:<17}{:<21}{:<45}{:<17} | '.format('Date', 'Diagnosis', 'Treatment', 'Term of treatment'))
    print('-'*104)

def med_hist_body(row):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    for dg, dt, tr, ds in sh['B'+str(row):'E'+str(row)]:
        dt = datetime.strftime(dt.value, '%d.%m.%Y')
        print('| {:<17}{:<21}{:<45}{:<17} | '.format(dt, dg.value, tr.value, f'{ds.value} days'))
        print('-'*104)

def medical_history(name, string='', header='Medical history'): 
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    names = []
    for i in sh['A']:
        if i.value == 'Name':
            continue
        if i.value not in names and i.value != None:
            names.append(i.value)
    if name not in names:
        print(string.center(104))
        return string
    else:
        first, second = True, True
        for i in sh['A']:
            if name == i.value:
                if first:
                    if header != '':
                        print(header.center(104))
                    med_hist_header()
                    first = False
                med_hist_body(i.row)

def last_date(name):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    dates = [sh['C'+str(i.row)].value for i in sh['A'] if name != 'Name' and name != None and name == i.value]
    last = max(dates)
    for j in sh['C']:
        if last == j.value and name == sh['A'+str(j.row)].value:
            row = j.row
    return last, row

def last_record(name):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    try:
        last, row = last_date(name)
        dt = datetime.strftime(last, '%d.%m.%Y')
        dg = sh['B'+str(row)].value
        tr = sh['D'+str(row)].value
        ds = sh['E'+str(row)].value
        print('Last record'.center(104))
        print('-'*104)
        print('| {:<17}{:<21}{:<45}{:<17} | '.format('Date', 'Diagnoses', 'Treatment', 'Term of treatment'))
        print('-'*104)
        print('| {:<17}{:<21}{:<45}{:<17} | '.format(dt, dg, tr, f'{ds} days'))
        print('-'*104)
    except:
        print('Your medical history is empty. It will be filled by your doctor'.center(104))

def count_remainder(row, last):
    wb = xl.load_workbook('patients.xlsx')
    sh = wb['Diagnosis']
    day =  sh['E'+str(row)].value
    delta = datetime.now() - last
    remainder = int(day) - delta.days
    return remainder, day

def treatment(name):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    try:
        last, row = last_date(name)
        remainder, day = count_remainder(row, last)
        if remainder > 0:
            print(f"The treatment duration - {day} days, {remainder} days left until the end".center(104))
        else:
            print(f"You have no treatment at the moment".center(104))
    except:
        print('Your medical history is empty'.center(104))

def schedule(): 
    wb = xl.load_workbook(path + 'doctors.xlsx')
    sh = wb['Schedule']
    print('Doctors schedule'.center(104))
    header = [sh.cell(row=1, column=i).value for i in range(1, 7)]
    table = PrettyTable(header)
    for a, b, c, d, e, f in sh['A2':'F10']:
        if a.value == None:
            a.value = ''
        values = [a.value, b.value, c.value, d.value, e.value, f.value]
        if values == [''] + [None for i in range(5)]:
            values = [''] + ['' for i in range(5)]
        table.add_row(values)
    print(table)

def patient_info(row, string, whose, header):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Logins']
    info = [sh.cell(row=row, column=i).value for i in range(4, 8)]
    try:
        info[2] = info[2].strftime('%d.%m.%Y')
        print(header.center(104))
        table = PrettyTable(['Height', 'Weight', 'Birthday date', 'Blood group'])
        table.add_row(['{:^20}'.format(info[0]), '{:^20}'.format(info[1]), '{:^30}'.format(info[2]), '{:^30}'.format(info[3])])
        print(table)
    except:
        print(string.center(104))
        hgt = float(input(f'Enter {whose} height: >>> '))
        wgt = float(input(f'Enter {whose} weight: >>> '))
        bth = input(f'Enter {whose} birthdate: >>> ')
        bth = datetime.strptime(bth, '%d.%m.%Y').date()
        bld = input(f'Enter {whose} bloodgroupe: >>> ')
        for h, w, bd, bg in sh['D'+str(row):'G'+str(row)]:
            h.value, w.value, bd.value, bg.value = hgt, wgt, bth, bld
        wb.save('patients.xlsx')
        patient_info(row, string, whose, header)

def patient():
    row, name = start('patients.xlsx')
    print('''    1 - show the medical history
    2 - show the last record in the medical history
    3 - show the number of days of treatment
    4 - show doctors' schedule
    5 - show my info
    6 - return to the main menu
    7 - exit''')
    while True:
        number = input('>>> ')
        if number == '1':
            s = 'Your medical history is empty. It will be filled by your doctor'
            medical_history(name, s)
        elif number == '2':
            last_record(name)
        elif number == '3':
            treatment(name)
        elif number == '4':
            schedule()
        elif number == '5':
            s = 'Your date is not yet in the system. Please fill in info about yourself'
            patient_info(row, s, 'your', 'My info')
        elif number == '6':
            break
        elif number == '7':
            print(number)
            exit('The program is over, we look forward to your return!'.center(104))
        else:
            print('Such command does not exists'.center(104))

def find_procedures(name, date):
    wb = xl.load_workbook(path + 'patients2.xlsx')
    sh = wb['Diagnosis']
    procedures = []
    for pr, start, end, days, time, med in sh['F2': 'K'+str(sh.max_row)]:
        if pr.value != None:
            if name == med.value:
                start, end = start.value.date(), end.value.date()
                if start <= date <= end:
                    days = days.value.split(', ')
                    day = calendar.day_name[date.weekday()]
                    if day in days:
                        patient_name = sh['A'+str(pr.row)].value
                        time = datetime.strftime(time.value, '%H:%M')
                        procedures.append((patient_name, pr.value, time))
    return procedures

def header(dt, day):
    print(f"Schedule of procedures {dt}".center(104))
    line = '-' * 76
    print(line.center(105))
    print('|{:^74}|'.center(38).format(day))
    print(line.center(105))
    print('| {:^3} | {:^10} | {:^25} | {:^25} |'.center(65).format('№', 'Time', 'Name', 'Procedure'))
    print(line.center(105))

def body(c, time, name, procedure):
    line = '-' * 76
    print('| {:^3} | {:^10} | {:^25} | {:^25} |'.center(65).format(c, time, name, procedure))
    print(line.center(105))

def procedures_table(name, dt, dt_obj):
    records = find_procedures(name, dt_obj)
    records.sort(key=lambda x: x[2])
    day = calendar.day_name[dt_obj.weekday()]
    header(dt, day)
    c = 1
    for record in records:
        name, procedure, time = record
        body(c, time, name, procedure) 
        c += 1

def procedures(name):
    wb = xl.load_workbook(path + 'patients2.xlsx')
    sh = wb['Diagnosis']
    print('For when do you want to see a schedule of procedures?'.center(104))
    print('1 - today, 2 - tomorrow, 3 - other date, 4 - go back'.center(104))
    while True:
        dt = input('Type (1/2/3/4) >>> ').strip()
        if dt == '1':
            today = date.today()
            today_str = datetime.strftime(today, '%d.%m.%Y')
            procedures_table(name, today_str, today)
        elif dt == '2':
            tomorrow = date.today()+timedelta(days=1)
            tomorrow_str = datetime.strftime(tomorrow, '%d.%m.%Y')
            procedures_table(name, tomorrow_str, tomorrow)
        elif dt == '3':
            dt = input('(dd.mm.yyyy) >>> ').strip()
            dt_obj = datetime.strptime(dt, '%d.%m.%Y').date()
            procedures_table(name, dt, dt_obj)
        elif dt == '4':
            break
        else:
            print('Wrong command'.center(104))

def write_errand(sheet, name, errand, dt):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    if sheet == 'Executed':
        date = input(f'Type the date of executing: >>> ')
        date = datetime.strptime(date, '%d.%m.%Y')
    else:
        date = dt
    r = sh['D1'].value
    for dt, er, md in sh['A'+str(r+1): 'C'+str(r+1)]:
        dt.value, er.value, md.value = date, errand, name
        sh['D1'].value += 1
    wb.save('errands.xlsx')
    if sheet == 'Executed':
        print('The list of executed errands'.center(104))
    errands_table('Executed', 'By')

def check_errand(name, string):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb['Appointed']
    errand = input(string).strip()
    for i, j in zip(sh['B'], sh['C']):
        if errand == i.value and name == j.value:
            break
    else:
        while errand != i.value or name != j.value:
            print("You don't have such errand, try again".center(104))
            errand = input(string).strip()
            for i, j in zip(sh['B'], sh['C']):
                if errand == i.value and name == j.value:
                    break
    return errand, i.row, sh['A'+str(i.row)].value

def count_errands(name):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb['Appointed']
    counter = 0
    for medass in sh['C']:
        if name == medass.value:
            counter += 1
    return counter

def execute_errand(name):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb['Appointed']
    counter = count_errands(name)
    if counter > 0:
        print(f'You have {counter} errand(-s)'.center(104))
        answer = input('Do you want to execute it/them? (y/n) >>> ').strip()
        if answer == 'y':
            string = 'Type an errand you wanna execute: >>> '
            errand, row, date = check_errand(name, string)
            sh.delete_rows(row)
            sh['D1'].value -= 1
            wb.save('errands.xlsx')
            write_errand('Executed', name, errand, date)
            medass_errands_menu(name, errand, date)
    else:
        print('You have no errands at the moment'.center(104))

def cancel_execution(name, errand, date):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb['Executed']
    for i in sh['B']:
        if errand == i.value:
            sh.delete_rows(i.row)
            sh['D1'].value -= 1
            wb.save('errands.xlsx')
            print('The execution has been cancelled'.center(104))
            write_errand('Appointed', name, errand, date)
        
def edit_date(sheet, errand):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    for i in sh['B']:
        if errand == i.value:
            date = input('Type a new date: >>> ').strip()
            date = datetime.strptime(date, '%d.%m.%Y')
            sh['A'+str(i.row)].value = date
            wb.save('errands.xlsx')
            print('The date has been edited'.center(104))
            show_certain_errand(sheet, errand)

def medass_errands_menu(name, errand, date):
    print('1 - edit date of executing, 2 - cancel execution, 3 - go back'.center(104))
    while True:
        option = input('Type (1/2/3) >>> ').strip()
        if option == '1':
            edit_date('Executed', errand)
        elif option == '2':
            print("Do you want to cancel the execution of the errand you've just typed?".center(104))
            if input('(y/n) >>> ').strip() == 'y':
                cancel_execution(name, errand, date)
                break
        elif option == '3':
            break
        else:
            print('Wrong command'.center(104))

def medassistant():
    row, name = start('medassistants.xlsx')
    print('''    1 - show a list of procedures
    2 - search for a patient
    3 - show a list of medassistants' errands
    4 - execute an errand
    5 - show executed errands
    6 - return to the main menu
    7 - exit''')
    while True:
        number = input('>>> ')
        if number == '1':
            procedures(name)
        elif number == '2':
            patient = input("Type a patient's name: >>> ")
            r = find_patient(patient)
            if r == 'No results':
                print(r.center(104))
        elif number == '3':
            print('The list of appointed errands'.center(104))
            errands_table('Appointed', 'For')
        elif number == '4':
            execute_errand(name)
        elif number == '5':
            print('The list of executed errands'.center(104))
            errands_table('Executed', 'By')
        elif number == '6':
            break
        elif number == '7':
            exit('The program is over, we look forward to your return!'.center(104))
        else:
            print('Such command does not exists'.center(104))

def patients_receiving_treatment():
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Logins']
    print('The list of patients receiving treatment'.center(104))
    line = '-'*49
    print(line.center(104))
    print('| {:<25}{:<10}{:>10} |'.center(75).format('Names', 'Days', 'Left'))
    print(line.center(104))
    for j in sh['A']:
        try: 
            last, row = last_date(j.value)
            remainder, day = count_remainder(row, last)
        except:
            continue
        if remainder > 0:
            print('| {:<25}{:<10}{:>10} |'.center(75).format(j.value, day, remainder))
    print(line.center(104))

def patients_quantity():
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Logins']
    c = 1
    while sh.cell(row=c, column=1).value!=None:
        c+=1
    print('Total number of patients'.center(104))
    print('+------+'.center(104))
    print(f' |  {c-2}  |'.center(104))
    print('+------+'.center(104))

def delete_errand(sheet):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    errand = input('Type an errand you wanna delete: >>> ').strip()
    for i in sh['B']:
        if errand == i.value:
            sh.delete_rows(i.row)
            sh['D1'].value -= 1 
            wb.save('errands.xlsx')
            print('The errand has been removed'.center(104))
            if sheet == 'Appointed':
                errands_table(sheet, 'For')
            else:
                errands_table(sheet, 'By')
            break
    else:
        print('No such errand'.center(104))

def show_certain_errand(sheet, errand):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    line = '-'*104
    print(line.center(110))
    c = 0
    for dt, er, medass in sh['A2': 'C'+str(sh.max_row)]:
        if dt.value != None and dt.value != 'Date':
            dt = datetime.strftime(dt.value, '%d.%m.%Y')
            c += 1
            if er.value == errand:
                print('|  {:<3} |  {:<19} |  {:<45} |  {:<20} |'.center(45).format(c, dt, er.value, medass.value))
    print(line.center(110))

def edit_errand(sheet):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    errand_is_changed = False
    errand = input('Type an errand you wanna edit: >>> ').strip()
    for i in sh['B']:
        if errand == i.value:
            print('What do you want to edit?'.center(104))
            print(f'1 - date, 2 - errand, 3 - name of medassistant, 4 - go back'.center(104))
            while True:
                option = input('(1/2/3/4) >>> ').strip()
                if option == '1':
                    date = input('Type a new date: >>> ').strip()
                    sh['A'+str(i.row)].value = datetime.strptime(date, '%d.%m.%Y')
                    wb.save('errands.xlsx')
                    print('The date has been edited'.center(104))
                    if errand_is_changed:
                        show_certain_errand(sheet, new_errand)
                    else:
                        show_certain_errand(sheet, errand)
                elif option == '2':
                    new_errand = input('Type a new errand: >>> ').strip()
                    sh['B'+str(i.row)].value = new_errand
                    wb.save('errands.xlsx')
                    errand_is_changed = True
                    print('The errand has been edited'.center(104))
                    show_certain_errand(sheet, new_errand)
                elif option == '3':
                    name = check_name('Type a new name: >>> ').strip()
                    sh['C'+str(i.row)].value = name
                    wb.save('errands.xlsx')
                    print('The name has been edited'.center(104))
                    if errand_is_changed:
                        show_certain_errand(sheet, new_errand)
                    else:
                        show_certain_errand(sheet, errand)
                elif option == '4':
                    break
                else:
                    print('Wrong command'.center(104))

def errands_table(sheet, preposition):
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb[sheet]
    line = '-'*104
    print(line.center(110))
    if sheet == 'Appointed':
        print('|  {:<3} |  {:<19} |  {:<45} |  {:<20} |'.center(45).format('№', 'Date of appointing', f'{sheet} errands', f'{preposition} whom'))
    else:
        print('|  {:<3} |  {:<19} |  {:<45} |  {:<20} |'.center(45).format('№', 'Date of executing', f'{sheet} errands', f'{preposition} whom'))
    print(line.center(110))
    c = 1
    for dt, er, medass in sh['A2': 'C'+str(sh.max_row)]:
        if dt.value != None and dt.value != 'Date':
            dt = datetime.strftime(dt.value, '%d.%m.%Y')
            print('|  {:<3} |  {:<19} |  {:<45} |  {:<20} |'.center(45).format(c, dt, er.value, medass.value))
            c += 1
    print(line.center(110))

def check_name(string):
    wb2 = xl.load_workbook(path+'medassistants.xlsx')
    sh2 = wb2['Logins']
    medass = input(string).strip()
    for i in sh2['A']:
        if medass == i.value:
            break
    else:
        while medass != i.value:
            print('No such medassistant, try again'.center(104))
            medass = input(string).strip()
            for i in sh2['A']:
                if medass == i.value:
                    break
    return medass 

def appoint_errand():
    wb = xl.load_workbook(path+'errands.xlsx')
    sh = wb['Appointed']
    errand = input('Type an errand: >>> ').strip()
    date = input(f'Type the date of appointing: >>> ')
    date = datetime.strptime(date, '%d.%m.%Y')
    medass = check_name('For whom: >>> ')
    r = sh['D1'].value
    for dt, er, md in sh['A'+str(r+1): 'C'+str(r+1)]:
        dt.value, er.value, md.value = date, errand, medass
        sh['D1'].value += 1
    wb.save('errands.xlsx')
    print(f'The list of appointed errands'.center(104))
    errands_table('Appointed', 'For')

def errands_menu(file):
    print('1 - edit an errand, 2 - remove an errand, 3 - go back'.center(104))
    while True:
        option = input('Type (1/2/3) >>> ')
        if option == '1':
            edit_errand(file)
        elif option == '2':
            delete_errand(file)
        elif option == '3':
            break
        else:
            print('Wrong command'.center(104))

def find_patient(name):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh1 = wb['Logins']
    sh2 = wb['Diagnosis']
    for i in sh1['A']:
        if name == i.value:
            try:
                last, row = last_date(name)
                dt = datetime.strftime(last, '%d.%m.%Y')
                dg = sh2['B'+str(row)].value
                line = '-'*71
                print(line.center(111))
                print('| {:^20}|{:^20}|{:^25} | '.center(65).format(name, dt, dg))
                print(line.center(111))
            except:
                line = '-'*70
                print(line.center(111))
                print('| {:^20}|{:^45} | '.center(60).format(name, 'any diagnosis'))
                print(line.center(111))
            return i.row
    else:
        return 'No results'

def searching():
    patient = input("Type a patient's name: >>> ")
    r = find_patient(patient)
    if r == 'No results':
        print(r.center(104))
    else:
        print("1 - patient's info, 2 - medical history, 3 - go back".center(104))
        while True:
            n = input('Type (1/2/3) >>> ')
            if n == '1':
                s1 = f"{patient}'s data is not yet in the system. Please fill in info about him/her".center(104)
                patient_info(r, s1, "patient's", f"{patient}'s info")
            elif n == '2':
                s2 = f"{patient}'s medical history is empty. Would you like to fill in it?".center(104)
                result = medical_history(patient, s2, f"{patient}'s medical history")
                if result == s2:
                    if input('(y/n) >>> ') == 'y':
                        write_diagnosis(patient)
                else:
                    med_hist_menu(patient)
            elif n == '3':
                break
            else:
                print('Wrong command'.center(104))

def maxx_row():
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Diagnosis']
    c = 1
    while sh.cell(row=c, column=1).value!=None:
        c+=1
    return c

def write_diagnosis(name):
    wb = xl.load_workbook('patients.xlsx')
    sh = wb['Diagnosis']
    dg = input('Enter a diagnosis for the patient: >>> ').capitalize()
    f = open('diagnosis.txt')
    for i in f:
        diagnosis, treatment = i.split(': ')
        if diagnosis == dg:
            dt = input('Enter the date of diagnosing: >>> ')
            dt = datetime.strptime(dt, '%d.%m.%Y')
            prescription, term = treatment.split(', ')
            num, word = term.split(' ')
            row = maxx_row()
            for a, b, c, d, e in sh['A'+str(row): 'E'+str(row)]:
                a.value, b.value, c.value, d.value, e.value = name, dg, dt, prescription, int(num)
                wb.save('patients.xlsx')
                medical_history(name, header=f"{name}'s medical history")
                med_hist_menu(name)
            break
    else:
        print('No such diagnosis'.center(104))
        
def diagnose(name):
    wb = xl.load_workbook(path + 'patients.xlsx')
    sh = wb['Logins']
    for i in sh['A']:
        if name == i.value:
            try:
                last, row = last_date(name)
                remainder, day = count_remainder(row, last)
                if remainder > 0:
                    print('The patient is already receiving treatment'.center(104))
                    med_hist_menu(name)
                else:
                    write_diagnosis(name)                
            except:
                write_diagnosis(name)
            break
    else:
        print('There is no such patient'.center(104))   

def new_diagnosis(row):
    wb = xl.load_workbook('patients.xlsx')
    sh = wb['Diagnosis']
    dg = input('Enter a new diagnosis: >>> ').capitalize().strip()
    f = open('diagnosis.txt')
    for i in f:
        diagnosis, treatment = i.split(': ')
        if diagnosis == dg:
            prescription, term = treatment.split(', ')
            num, word = term.split(' ')
            sh['B'+str(row)].value, sh['D'+str(row)].value, sh['E'+str(row)].value  = dg, prescription, int(num)
            wb.save('patients.xlsx')
            print('The diagnosis has been edited'.center(104))
            break
    else:
        print('No such diagnosis'.center(104))

def edit_diagnosis(name):
    wb = xl.load_workbook('patients.xlsx')
    sh = wb['Diagnosis']
    print('Which record do you want to edit?'.center(104))
    record = input('Input date: >>> ').strip()
    record = datetime.strptime(record, '%d.%m.%Y')
    for i, j in zip(sh['A'], sh['C']):
        if name == i.value and record == j.value:
            print('Which data?'.center(104))
            print('1 - diagnosis, 2 - date of diagnosing, 3 - treatment, 4 - term of treatment'.center(104))
            data = input('(1/2/3/4) >>> ').strip()
            if data == '1':
                new_diagnosis(i.row)
            elif data == '2':
                dt = input('Enter a new date: >>> ').strip()
                dt = datetime.strptime(dt, '%d.%m.%Y')
                sh['C'+str(i.row)] = dt
                wb.save('patients.xlsx')
                print('The date has been edited'.center(104))
            elif data == '3':
                sh['D'+str(i.row)] = input('Enter a new treatment: >>> ').strip()
                wb.save('patients.xlsx')
                print('The treatment has been edited'.center(104))
            elif data == '4':
                sh['E'+str(i.row)] = int(input('Enter a new term of treatment: >>> '))
                wb.save('patients.xlsx')
                print('The term of treatment has been edited'.center(104))
            medical_history(name, header='')
            break
    else:
        print('No such record'.center(104))

def delete_diagnosis(name):
    wb = xl.load_workbook('patients.xlsx')
    sh = wb['Diagnosis']
    print('Which record do you want to delete?'.center(104))
    record = input('Input date: >>> ').strip()
    record = datetime.strptime(record, '%d.%m.%Y')
    for i, j in zip(sh['A'], sh['C']):
        if name == i.value and record == j.value:
            print('This record would be deleted'.center(104))
            if input('Continue? (y/n) >>> ') == 'y':
                sh.delete_rows(i.row)
                wb.save('patients.xlsx')
                print('The record has been deleted'.center(104))
                s = f"{name}'s medical history is empty".center(104)
                medical_history(name, s, '')
            break
    else:
        print('No such record'.center(104))

def med_hist_menu(name):
    print('1 - edit diagnosis, 2 - delete diagnosis, 3 - go back'.center(104))
    while True:
        option = input('(1/2/3) >>> ')
        if option == '1':
            edit_diagnosis(name)
        elif option == '2':
            delete_diagnosis(name)
        elif option == '3':
            break
        else:
            print('Wrong command'.center(104))

def doctor():
    row, name = start('doctors.xlsx')
    print('''    1 - show a list of patients receiving treatment
    2 - show the total number of patients
    3 - show a list of medassistants' errands
    4 - write an errand for a medassistant
    5 - show executed errands
    6 - search for a patient
    7 - diagnose a patient
    8 - return to the main menu
    9 - exit''')
    while True:
        number = input('>>> ')
        if number == '1':
            patients_receiving_treatment()
        elif number == '2':
            patients_quantity()
        elif number == '3': 
            print('The list of appointed errands'.center(104))
            errands_table('Appointed', 'For')
            errands_menu('Appointed')
        elif number == '4':
            appoint_errand()
            errands_menu('Appointed')
        elif number == '5': 
            print('The list of executed errands'.center(104))
            errands_table('Executed', 'By')
            errands_menu('Executed')
        elif number == '6':
            searching()
        elif number == '7':
            patient = input("Type a patient's name: >>> ").strip()
            diagnose(patient)
        elif number == '8':
            break
        elif number == '9':
            exit('The program is over, we look forward to your return!'.center(104))
        else:
            print('Such command does not exists'.center(104))

def maindoctor():
    row, name = start('maindoctors.xlsx')
    print('''    1 - show the list of medassistances
    2 - show the list of doctors
    3 - show amount of patients 
    4 - show the employee with the highest salary
    5 - show the employee with the lowest salary
    6 - return to the main menu
    7 - exit''')


commands = {
    'patient': patient,
    'medassistant': medassistant,
    'doctor': doctor,
    'maindoctor': maindoctor
}    

print('Welcome to the information system "AIS Hospital"'.center(104))
while True:
    print('''Enter your account type:
    -patient
    -medassistant
    -doctor
    -maindoctor
    -exit''')
    command = input('>>>> ')  
    if command in commands:
        commands[command]()
    elif command == 'exit':
        exit('The program is over, we look forward to your return!'.center(104))
    else:
        print('Wrong account type, please enter again'.center(104))
